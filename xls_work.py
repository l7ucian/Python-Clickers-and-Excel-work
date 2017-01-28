from openpyxl import load_workbook
wb = load_workbook(filename = '')         

#grab worksheet by name (use active for active worksheet)
#G la K := luni pana vineri
ws = wb.get_sheet_by_name("")

days = {
    'G': 'Luni',
    'H': 'Marti',
    'I': 'Miercuri',
    'J': 'Joi',
    'K': 'Vineri'
}

list = "BCDEFGHIJK"

# ws['D7'].number_format = '0%'
# ws['D7'] = '100%'
#ws['K9'] = 6

for i in range(5,16):
    for letter in list:
        if letter in "GHIJK":
            print(letter + str(i)," "+days[letter]+ " Ore muncite: ",ws[letter+str(i)].value)
        else:
            print(letter+str(i),ws[letter+str(i)].value)


    print('-----------------------------------------')
    print()
#wb.save(')
